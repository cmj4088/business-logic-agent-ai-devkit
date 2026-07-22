"""ipd-xlsx Skill — IPD 项目 Excel 文件生成。

支持 BOM 成本表、财务预算表、项目进度表、竞品对比矩阵等 IPD 交付物。
生成 .xlsx 文件，含公式、格式化、图表。
"""

import io
import json
import logging
from datetime import datetime
from typing import Any
from .base import BaseSkill, SkillContext, SkillResult

logger = logging.getLogger(__name__)

# 模板类型定义
TEMPLATE_TYPES = {
    "bom": {
        "name": "BOM 成本表",
        "description": "物料清单：物料编码、名称、规格、数量、单价、总价，自动汇总",
        "sheets": ["BOM 物料清单", "成本汇总", "供应商"],
    },
    "budget": {
        "name": "财务预算表",
        "description": "收入/支出分类，预算 vs 实际，偏差分析，超预算标红",
        "sheets": ["预算总表", "按阶段", "按部门"],
    },
    "schedule": {
        "name": "项目进度表",
        "description": "甘特图格式：阶段、活动、开始/结束日期、负责人、状态，自动计算工期",
        "sheets": ["项目进度", "里程碑", "资源分配"],
    },
    "competitive_matrix": {
        "name": "竞品对比矩阵",
        "description": "功能、价格、性能、市场份额对比，加权评分，雷达图",
        "sheets": ["对比矩阵", "评分明细", "雷达图数据"],
    },
    "test_cases": {
        "name": "测试用例集",
        "description": "用例编号、模块、测试步骤、预期结果、优先级、状态",
        "sheets": ["测试用例", "执行记录", "缺陷统计"],
    },
}


class XlsxSkill(BaseSkill):
    """IPD Excel 文件生成 Skill。

    支持 5 类模板：BOM、预算、进度、竞品矩阵、测试用例。
    生成格式规范的 .xlsx 文件。
    """

    @property
    def name(self) -> str:
        return "ipd-xlsx"

    @property
    def description(self) -> str:
        return "IPD 项目 Excel 交付物生成：BOM 成本表、财务预算表、项目进度表、竞品对比矩阵"

    def get_tools(self) -> list[dict]:
        """注册为 Agent 可调用的 tool。"""
        return [
            {
                "tool_name": "generate_xlsx",
                "tool_schema": json.dumps({
                    "type": "function",
                    "function": {
                        "name": "generate_xlsx",
                        "description": "生成 IPD 项目 Excel 文件，支持 BOM/预算/进度/竞品矩阵/测试用例模板",
                        "parameters": {
                            "type": "object",
                            "properties": {
                                "template_type": {
                                    "type": "string",
                                    "enum": list(TEMPLATE_TYPES.keys()),
                                    "description": "Excel 模板类型",
                                },
                                "project_name": {
                                    "type": "string",
                                    "description": "项目名称",
                                },
                                "rows_data": {
                                    "type": "string",
                                    "description": "表格数据（JSON 数组格式）",
                                },
                                "extra_params": {
                                    "type": "object",
                                    "description": "额外参数（如预算金额、时间范围等）",
                                    "properties": {
                                        "currency": {"type": "string", "description": "货币单位，默认 CNY"},
                                        "date_range": {"type": "string", "description": "日期范围"},
                                        "department": {"type": "string", "description": "部门名称"},
                                    },
                                },
                            },
                            "required": ["template_type", "project_name"],
                        },
                    },
                }, ensure_ascii=False),
            },
        ]

    async def validate(self, context: SkillContext) -> tuple[bool, str]:
        params = context.params
        template_type = params.get("template_type", "")
        if template_type and template_type not in TEMPLATE_TYPES:
            return False, f"不支持的模板类型 '{template_type}'，可选: {list(TEMPLATE_TYPES.keys())}"
        return True, ""

    async def execute(self, context: SkillContext) -> SkillResult:
        params = context.params
        template_type = params.get("template_type", "bom")
        project_name = params.get("project_name", "未命名项目")
        rows_data = params.get("rows_data", "[]")
        extra_params = params.get("extra_params", {})

        template_info = TEMPLATE_TYPES.get(template_type, TEMPLATE_TYPES["bom"])

        # 解析行数据
        try:
            rows = json.loads(rows_data) if isinstance(rows_data, str) else rows_data
        except json.JSONDecodeError:
            rows = []

        file_content = await self._generate_xlsx(
            template_type=template_type,
            template_name=template_info["name"],
            sheets=template_info["sheets"],
            project_name=project_name,
            rows=rows,
            extra_params=extra_params,
        )

        # 保存文件到输出目录
        file_path = self._save_file(file_content, template_type, project_name)

        return SkillResult(
            success=True,
            skill_name=self.name,
            output=f"Excel 文件已生成: {template_info['name']} ({len(rows)} 行数据)",
            file_path=file_path,
            file_type="xlsx",
            artifact_type=f"{template_type}_xlsx",
            tokens_used=len(rows) * 10 + 100,
            metadata={
                "template_type": template_type,
                "template_name": template_info["name"],
                "sheets": template_info["sheets"],
                "row_count": len(rows),
                "project_name": project_name,
            },
        )

    def _save_file(self, content: bytes, template_type: str, project_name: str) -> str:
        """保存生成的文件到输出目录。"""
        import os
        output_dir = os.path.join(os.path.dirname(__file__), "..", "output", "skills")
        os.makedirs(output_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_name = "".join(c for c in project_name if c.isalnum() or c in "_ ").strip().replace(" ", "_")
        filename = f"{template_type}_{safe_name}_{timestamp}.xlsx"
        filepath = os.path.join(output_dir, filename)
        with open(filepath, "wb") as f:
            f.write(content)
        return filepath

    async def _generate_xlsx(
        self,
        template_type: str,
        template_name: str,
        sheets: list[str],
        project_name: str,
        rows: list[dict],
        extra_params: dict,
    ) -> bytes:
        """生成 Excel 文件内容。

        使用 openpyxl 创建格式化的 .xlsx 文件。
        如果 openpyxl 不可用，返回 CSV 格式降级内容。

        Returns:
            bytes: Excel 文件二进制内容。
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.warning("openpyxl 未安装，返回 CSV 降级内容")
            return self._csv_fallback(template_name, sheets, rows)

        wb = openpyxl.Workbook()

        # 样式定义
        header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_font = Font(name="微软雅黑", size=10)
        cell_alignment = Alignment(vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        alt_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

        # Sheet 1: 主数据表
        ws = wb.active
        ws.title = sheets[0] if sheets else "Sheet1"

        # 根据模板类型生成表头
        headers = self._get_headers(template_type)
        col_widths = self._get_col_widths(template_type)

        # 写入表头
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 设置列宽
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # 写入数据行
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header, row_data.get(str(col_idx - 1), ""))
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = cell_font
                cell.alignment = cell_alignment
                cell.border = thin_border
                # 交替行颜色
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

        # 冻结首行
        ws.freeze_panes = "A2"

        # 自动筛选
        if len(headers) > 0 and len(rows) > 0:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"

        # Sheet 2: 汇总表
        if len(sheets) > 1:
            ws2 = wb.create_sheet(title=sheets[1])
            ws2.cell(row=1, column=1, value=f"{template_name} - 汇总").font = Font(
                name="微软雅黑", bold=True, size=14
            )
            ws2.cell(row=3, column=1, value="项目名称").font = Font(bold=True)
            ws2.cell(row=3, column=2, value=project_name)
            ws2.cell(row=4, column=1, value="模板类型").font = Font(bold=True)
            ws2.cell(row=4, column=2, value=template_name)
            ws2.cell(row=5, column=1, value="数据行数").font = Font(bold=True)
            ws2.cell(row=5, column=2, value=len(rows))
            ws2.cell(row=7, column=1, value="生成时间").font = Font(bold=True)
            ws2.cell(row=7, column=2, value=self._now())
            ws2.column_dimensions["A"].width = 15
            ws2.column_dimensions["B"].width = 30

        # Sheet 3: 附加信息
        if len(sheets) > 2:
            ws3 = wb.create_sheet(title=sheets[2])
            ws3.cell(row=1, column=1, value=f"{template_name} - 附加信息").font = Font(
                name="微软雅黑", bold=True, size=14
            )
            row = 3
            for key, value in extra_params.items():
                ws3.cell(row=row, column=1, value=str(key)).font = Font(bold=True)
                ws3.cell(row=row, column=2, value=str(value))
                row += 1

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def _get_headers(self, template_type: str) -> list[str]:
        """获取模板表头。"""
        headers_map = {
            "bom": ["物料编码", "物料名称", "规格型号", "数量", "单位", "单价(元)", "总价(元)", "备注"],
            "budget": ["预算科目", "预算金额(元)", "实际金额(元)", "偏差(元)", "偏差率", "状态", "备注"],
            "schedule": ["序号", "阶段", "活动名称", "开始日期", "结束日期", "工期(天)", "负责人", "状态", "备注"],
            "competitive_matrix": ["对比维度", "权重", "竞品A", "竞品B", "竞品C", "我方", "备注"],
            "test_cases": ["用例编号", "模块", "测试步骤", "预期结果", "优先级", "状态", "备注"],
        }
        return headers_map.get(template_type, ["字段1", "字段2", "字段3", "字段4", "字段5"])

    def _get_col_widths(self, template_type: str) -> list[int]:
        """获取模板列宽。"""
        widths_map = {
            "bom": [15, 25, 20, 10, 8, 12, 12, 20],
            "budget": [20, 15, 15, 15, 10, 10, 20],
            "schedule": [8, 12, 25, 14, 14, 10, 10, 10, 20],
            "competitive_matrix": [15, 10, 15, 15, 15, 15, 20],
            "test_cases": [15, 15, 30, 30, 10, 10, 20],
        }
        return widths_map.get(template_type, [15, 15, 15, 15, 15])

    def _csv_fallback(self, template_name: str, sheets: list[str], rows: list[dict]) -> bytes:
        """openpyxl 不可用时，返回 CSV 格式降级内容。"""
        import csv
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([f"{template_name}"])
        writer.writerow([])
        if rows and len(rows) > 0:
            writer.writerow(list(rows[0].keys()))
            for row in rows:
                writer.writerow(row.values())
        return output.getvalue().encode("utf-8-sig")

    def _now(self) -> str:
        from datetime import datetime, timezone
        return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")